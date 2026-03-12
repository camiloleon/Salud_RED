from __future__ import annotations

import argparse
import hashlib
import os
import re
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Iterable

import mysql.connector
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from mysql.connector.connection import MySQLConnection
from mysql.connector.abstracts import MySQLConnectionAbstract

BATCH_SIZE = 1000

DATE_CANDIDATES = [
    "fecha",
    "m_fecha_mes",
    "fech_ini_incid",
    "fech_fin_incid",
    "fecha_inicio_afectacion",
    "fecha_fin_afectacion",
]


def normalize_name(value: str, max_len: int = 60) -> str:
    text = value.strip().lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = "col"
    return text[:max_len]


def dedupe_columns(columns: Iterable[str]) -> list[str]:
    out = []
    seen: dict[str, int] = {}
    for c in columns:
        base = normalize_name(str(c), max_len=50)
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
    return out


def row_hash(values: list[str], sheet: str) -> str:
    payload = "||".join([sheet, *values])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


@dataclass
class SyncStats:
    source: str
    rows_read: int
    rows_loaded_raw: int
    rows_loaded_current: int


def get_mysql_connection(args: argparse.Namespace) -> MySQLConnectionAbstract:
    env_file = Path(args.env_file)
    if env_file.exists():
        load_dotenv(env_file)

    host = args.mysql_host or os.getenv("MYSQL_HOST")
    port = int(args.mysql_port or os.getenv("MYSQL_PORT", "3306"))
    user = args.mysql_user or os.getenv("MYSQL_USER")
    password = args.mysql_password or os.getenv("MYSQL_PASSWORD")
    database = args.mysql_db or os.getenv("MYSQL_DATABASE")

    missing = [
        name
        for name, val in [
            ("MYSQL_HOST", host),
            ("MYSQL_USER", user),
            ("MYSQL_PASSWORD", password),
            ("MYSQL_DATABASE", database),
        ]
        if not val
    ]
    if missing:
        raise ValueError(f"Missing MySQL settings: {', '.join(missing)}")

    return mysql.connector.connect(
        host=host,
        port=port,
        user=user,
        password=password,
        database=database,
        autocommit=False,
    )


def extract_excel_tables(workbook: Path, only_tables: set[str] | None = None) -> list[tuple[str, str, pd.DataFrame]]:
    wb = load_workbook(workbook, data_only=True, read_only=False)
    extracted: list[tuple[str, str, pd.DataFrame]] = []

    for ws in wb.worksheets:
        for table_name, table_obj in ws.tables.items():
            if only_tables and table_name not in only_tables:
                continue

            table_ref = table_obj.ref if hasattr(table_obj, "ref") else str(table_obj)
            data_cells = ws[table_ref]
            rows = [[cell.value for cell in row] for row in data_cells]
            if len(rows) < 2:
                continue

            headers = dedupe_columns([str(c) for c in rows[0]])
            body = rows[1:]
            df = pd.DataFrame(body, columns=headers)
            df = df.dropna(how="all")
            if df.empty:
                continue

            extracted.append((ws.title, table_name, df.fillna("")))

    return extracted


def ensure_table(conn: MySQLConnectionAbstract, table_name: str, columns: list[str]) -> None:
    cols_sql = ",\n  ".join([f"`{c}` TEXT NULL" for c in columns])
    sql = f"""
CREATE TABLE IF NOT EXISTS `{table_name}` (
  `row_hash` CHAR(64) NOT NULL,
  `source_sheet` VARCHAR(128) NOT NULL,
  `source_file` VARCHAR(255) NOT NULL,
  `ingested_at` DATETIME NOT NULL,
  {cols_sql},
  PRIMARY KEY (`row_hash`),
  KEY `idx_ingested_at` (`ingested_at`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
"""
    cur = conn.cursor()
    cur.execute(sql)
    cur.close()


def ensure_current_table(conn: MySQLConnectionAbstract, table_name: str, columns: list[str]) -> None:
    cols_sql = ",\n  ".join([f"`{c}` TEXT NULL" for c in columns])
    sql = f"""
CREATE TABLE IF NOT EXISTS `{table_name}` (
  `row_hash` CHAR(64) NOT NULL,
  `source_sheet` VARCHAR(128) NOT NULL,
  `source_file` VARCHAR(255) NOT NULL,
  `ingested_at` DATETIME NOT NULL,
  {cols_sql},
  PRIMARY KEY (`row_hash`),
  KEY `idx_ingested_at` (`ingested_at`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
"""
    cur = conn.cursor()
    cur.execute(sql)
    cur.close()


def detect_date_column(df: pd.DataFrame) -> str | None:
    lowered = {c.lower(): c for c in df.columns}
    for cand in DATE_CANDIDATES:
        if cand in lowered:
            return lowered[cand]
    for col in df.columns:
        if "fecha" in str(col).lower():
            return col
    return None


def filter_dataframe_by_mode(
    df: pd.DataFrame,
    mode: str,
    bootstrap_start: str,
    incremental_days: int,
) -> pd.DataFrame:
    date_col = detect_date_column(df)
    if not date_col:
        return df

    parsed = pd.to_datetime(df[date_col], errors="coerce")
    out = df.copy()
    out["__parsed_date"] = parsed

    if mode == "bootstrap":
        start = pd.to_datetime(bootstrap_start)
        out = out[out["__parsed_date"] >= start]
    else:
        start = pd.Timestamp(datetime.now(UTC).date() - timedelta(days=incremental_days))
        out = out[out["__parsed_date"] >= start]

    out = out.drop(columns=["__parsed_date"]) if "__parsed_date" in out.columns else out
    return out


def insert_raw(conn: MySQLConnectionAbstract, table_name: str, columns: list[str], records: list[dict]) -> int:
    if not records:
        return 0

    all_cols = ["row_hash", "source_sheet", "source_file", "ingested_at", *columns]
    placeholders = ",".join(["%s"] * len(all_cols))
    col_sql = ",".join([f"`{c}`" for c in all_cols])
    sql = f"INSERT INTO `{table_name}` ({col_sql}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE ingested_at=VALUES(ingested_at)"

    values = []
    for r in records:
        values.append([r.get(c) for c in all_cols])

    total = 0
    cur = conn.cursor()
    for i in range(0, len(values), BATCH_SIZE):
        chunk = values[i : i + BATCH_SIZE]
        cur.executemany(sql, chunk)
        total += int(cur.rowcount or 0)
    cur.close()
    return total


def replace_current(conn: MySQLConnectionAbstract, table_name: str, columns: list[str], records: list[dict]) -> int:
    cur = conn.cursor()
    cur.execute(f"TRUNCATE TABLE `{table_name}`")
    cur.close()

    if not records:
        return 0

    all_cols = ["row_hash", "source_sheet", "source_file", "ingested_at", *columns]
    placeholders = ",".join(["%s"] * len(all_cols))
    col_sql = ",".join([f"`{c}`" for c in all_cols])
    sql = f"INSERT INTO `{table_name}` ({col_sql}) VALUES ({placeholders})"

    values = []
    for r in records:
        values.append([r.get(c) for c in all_cols])

    total = 0
    cur = conn.cursor()
    for i in range(0, len(values), BATCH_SIZE):
        chunk = values[i : i + BATCH_SIZE]
        cur.executemany(sql, chunk)
        total += int(cur.rowcount or 0)
    cur.close()
    return total


def upsert_current(conn: MySQLConnectionAbstract, table_name: str, columns: list[str], records: list[dict]) -> int:
    if not records:
        return 0

    all_cols = ["row_hash", "source_sheet", "source_file", "ingested_at", *columns]
    placeholders = ",".join(["%s"] * len(all_cols))
    col_sql = ",".join([f"`{c}`" for c in all_cols])
    sql = f"INSERT INTO `{table_name}` ({col_sql}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE ingested_at=VALUES(ingested_at)"

    values = []
    for r in records:
        values.append([r.get(c) for c in all_cols])

    total = 0
    cur = conn.cursor()
    for i in range(0, len(values), BATCH_SIZE):
        chunk = values[i : i + BATCH_SIZE]
        cur.executemany(sql, chunk)
        total += int(cur.rowcount or 0)
    cur.close()
    return total


def sync_table(
    conn: MySQLConnectionAbstract | None,
    workbook: Path,
    sheet_name: str,
    table_name: str,
    df: pd.DataFrame,
    dry_run: bool,
    mode: str,
    bootstrap_start: str,
    incremental_days: int,
) -> SyncStats:
    df = filter_dataframe_by_mode(df, mode=mode, bootstrap_start=bootstrap_start, incremental_days=incremental_days)
    cols = list(df.columns)

    if dry_run:
        return SyncStats(source=f"{sheet_name}/{table_name}", rows_read=len(df), rows_loaded_raw=0, rows_loaded_current=0)

    now = datetime.now(UTC).replace(tzinfo=None, microsecond=0)

    records: list[dict[str, object]] = []
    for _, row in df.iterrows():
        row_vals = [str(row[c]) for c in cols]
        rec: dict[str, object] = {c: str(row[c]) for c in cols}
        rec["row_hash"] = row_hash(row_vals, f"{sheet_name}::{table_name}")
        rec["source_sheet"] = sheet_name
        rec["source_file"] = workbook.name
        rec["ingested_at"] = now
        records.append(rec)

    entity = normalize_name(f"{sheet_name}_{table_name}", max_len=48)
    raw_table = f"raw_{entity}"
    cur_table = f"cur_{entity}"

    assert conn is not None
    ensure_table(conn, raw_table, cols)
    ensure_current_table(conn, cur_table, cols)
    raw_rows = insert_raw(conn, raw_table, cols, records)
    if mode == "bootstrap":
        current_rows = replace_current(conn, cur_table, cols, records)
    else:
        current_rows = upsert_current(conn, cur_table, cols, records)
    return SyncStats(source=f"{sheet_name}/{table_name}", rows_read=len(df), rows_loaded_raw=raw_rows, rows_loaded_current=current_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Excel workbook tables to MySQL")
    parser.add_argument("--workbook", required=True, help="Path to workbook")
    parser.add_argument("--env-file", default=".env.mysql", help="Path to env file with MySQL credentials")
    parser.add_argument("--mysql-host")
    parser.add_argument("--mysql-port")
    parser.add_argument("--mysql-user")
    parser.add_argument("--mysql-password")
    parser.add_argument("--mysql-db")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only-tables", nargs="*", default=[])
    parser.add_argument("--mode", choices=["bootstrap", "incremental"], default="incremental")
    parser.add_argument("--bootstrap-start", default="2023-01-01")
    parser.add_argument("--incremental-days", type=int, default=2)
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    selected_tables = set(args.only_tables) if args.only_tables else None
    extracted = extract_excel_tables(workbook, only_tables=selected_tables)
    if not extracted:
        raise ValueError("No Excel tables were found to sync. Check table names and workbook.")

    if args.dry_run:
        print("[DRY-RUN] No DB connection will be opened. Tables detected:")
        for sheet_name, table_name, df in extracted:
            stats = sync_table(  # type: ignore[arg-type]
                None,
                workbook,
                sheet_name,
                table_name,
                df,
                dry_run=True,
                mode=args.mode,
                bootstrap_start=args.bootstrap_start,
                incremental_days=args.incremental_days,
            )
            print(f"source={stats.source} rows={stats.rows_read}")
        return 0

    conn = get_mysql_connection(args)
    try:
        total_raw = 0
        total_cur = 0
        for sheet_name, table_name, df in extracted:
            stats = sync_table(
                conn,
                workbook,
                sheet_name,
                table_name,
                df,
                dry_run=False,
                mode=args.mode,
                bootstrap_start=args.bootstrap_start,
                incremental_days=args.incremental_days,
            )
            total_raw += stats.rows_loaded_raw
            total_cur += stats.rows_loaded_current
            print(
                f"mode={args.mode} source={stats.source} rows_read={stats.rows_read} raw_upserted={stats.rows_loaded_raw} current_loaded={stats.rows_loaded_current}"
            )
        conn.commit()
        print(f"done raw_upserted_total={total_raw} current_loaded_total={total_cur}")
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
